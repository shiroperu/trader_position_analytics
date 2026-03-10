"""Phase 2: トークン別センチメント分析 & Excel出力"""

from __future__ import annotations

import glob
import os
import sys
from datetime import datetime
from pathlib import Path

# スタンドアロン実行時のパス解決
if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from scripts.config import DATA_DIR, STYLES, TRADERS, build_styles, setup_logging


# ============================================================
# ポジションExcel読み込み
# ============================================================

def load_positions(filepath: str) -> list[dict]:
    """ポジションExcelを読み込み、行データのリストを返す。"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    current_tier = ""
    current_trader = ""
    current_acct = 0.0

    for r in range(2, ws.max_row + 1):
        tier = ws.cell(r, 1).value
        trader = ws.cell(r, 2).value
        acct = ws.cell(r, 3).value
        coin = ws.cell(r, 4).value
        side = ws.cell(r, 5).value

        # グループ表示: Tier/Trader/Acctが空の場合は前の値を引き継ぐ
        if tier:
            current_tier = tier
        if trader:
            current_trader = trader
        if acct is not None:
            current_acct = float(acct)

        if coin == "(No positions)" or not coin:
            continue

        rows.append({
            "tier": current_tier,
            "trader": current_trader,
            "acct_value": current_acct,
            "coin": str(coin),
            "side": str(side) if side else "",
            "size": float(ws.cell(r, 6).value or 0),
            "leverage": float(ws.cell(r, 7).value or 0),
            "entry_px": float(ws.cell(r, 8).value or 0),
            "pos_value": float(ws.cell(r, 9).value or 0),
            "unrealized_pnl": float(ws.cell(r, 10).value or 0),
            "roe": float(ws.cell(r, 11).value or 0),
            "liq_px": ws.cell(r, 12).value,
            "margin": float(ws.cell(r, 13).value or 0),
            "cum_funding": float(ws.cell(r, 14).value or 0),
        })

    wb.close()
    return rows


def find_previous_file(current_file: str) -> str | None:
    """data/ 内の hl_positions_*.xlsx を日付順ソートし、current_fileの1つ前を返す。"""
    pattern = str(DATA_DIR / "hl_positions_*.xlsx")
    files = sorted(glob.glob(pattern))
    current_path = os.path.abspath(current_file)

    for i, f in enumerate(files):
        if os.path.abspath(f) == current_path and i > 0:
            return files[i - 1]
    return None


# ============================================================
# センチメント分析ロジック
# ============================================================

def analyze_tokens(rows: list[dict]) -> list[dict]:
    """トークン別のセンチメント集計を行う。"""
    token_data = {}
    for r in rows:
        coin = r["coin"]
        if coin not in token_data:
            token_data[coin] = {
                "long_traders": set(), "short_traders": set(),
                "long_value": 0.0, "short_value": 0.0,
                "long_pnl": 0.0, "short_pnl": 0.0,
                "long_leverages": [], "short_leverages": [],
            }
        td = token_data[coin]
        if r["side"] == "LONG":
            td["long_traders"].add(r["trader"])
            td["long_value"] += r["pos_value"]
            td["long_pnl"] += r["unrealized_pnl"]
            td["long_leverages"].append(r["leverage"])
        elif r["side"] == "SHORT":
            td["short_traders"].add(r["trader"])
            td["short_value"] += r["pos_value"]
            td["short_pnl"] += r["unrealized_pnl"]
            td["short_leverages"].append(r["leverage"])

    results = []
    for coin, td in token_data.items():
        long_count = len(td["long_traders"])
        short_count = len(td["short_traders"])
        total_count = long_count + short_count
        long_ratio = long_count / total_count if total_count > 0 else 0
        short_ratio = short_count / total_count if total_count > 0 else 0
        total_value = td["long_value"] + td["short_value"]
        net_direction = td["long_value"] - td["short_value"]
        total_pnl = td["long_pnl"] + td["short_pnl"]

        # 平均レバレッジ（L/S別）
        avg_lev_long = (sum(td["long_leverages"]) / len(td["long_leverages"])
                        if td["long_leverages"] else 0)
        avg_lev_short = (sum(td["short_leverages"]) / len(td["short_leverages"])
                         if td["short_leverages"] else 0)

        sentiment = _judge_sentiment(long_ratio, short_ratio)

        results.append({
            "coin": coin,
            "long_count": long_count,
            "short_count": short_count,
            "total_count": total_count,
            "long_ratio": long_ratio,
            "short_ratio": short_ratio,
            "sentiment": sentiment,
            "long_value": td["long_value"],
            "short_value": td["short_value"],
            "total_value": total_value,
            "net_direction": net_direction,
            "long_pnl": td["long_pnl"],
            "short_pnl": td["short_pnl"],
            "total_pnl": total_pnl,
            "avg_lev_long": avg_lev_long,
            "avg_lev_short": avg_lev_short,
        })

    # 合計価値の降順ソート
    results.sort(key=lambda x: x["total_value"], reverse=True)
    return results


def _judge_sentiment(long_ratio: float, short_ratio: float) -> str:
    if long_ratio >= 0.70:
        return "\U0001f7e2 Strong LONG"
    elif long_ratio >= 0.55:
        return "\u2197 Lean LONG"
    elif long_ratio == 0.50:
        return "\u2696 Neutral"
    elif short_ratio >= 0.70:
        return "\U0001f534 Strong SHORT"
    elif short_ratio >= 0.55:
        return "\u2198 Lean SHORT"
    else:
        return "\u2696 Neutral"


def compute_changes(current: list[dict], previous: list[dict]) -> dict:
    """前回データとの差分を算出。キー=coin, 値=変動指標dict。"""
    prev_map = {t["coin"]: t for t in previous}
    changes = {}
    for t in current:
        coin = t["coin"]
        if coin in prev_map:
            p = prev_map[coin]
            long_ratio_change = (t["long_ratio"] - p["long_ratio"]) * 100  # pt
            value_change_pct = (
                ((t["total_value"] - p["total_value"]) / p["total_value"] * 100)
                if p["total_value"] != 0 else 0
            )
            trader_count_change = t["total_count"] - p["total_count"]
            changes[coin] = {
                "long_ratio_change": long_ratio_change,
                "value_change_pct": value_change_pct,
                "trader_count_change": trader_count_change,
            }
    return changes


# ============================================================
# Trader×Token マトリクス構築
# ============================================================

def build_trader_matrix(rows: list[dict], top_coins: list[str]) -> tuple:
    """アクティブトレーダー × 上位トークンのマトリクスを構築。

    Returns: (traders_list, matrix_dict) where matrix_dict[(trader, coin)] = "L"/"S"
    """
    active_traders = []
    seen = set()
    for r in rows:
        key = r["trader"]
        if key not in seen:
            seen.add(key)
            active_traders.append({
                "trader": r["trader"],
                "tier": r["tier"],
                "acct_value": r["acct_value"],
            })

    matrix = {}
    for r in rows:
        if r["coin"] in top_coins:
            key = (r["trader"], r["coin"])
            matrix[key] = "L" if r["side"] == "LONG" else "S"

    return active_traders, matrix


# ============================================================
# Excel出力
# ============================================================

def _apply_header(ws, row, columns, widths, styles):
    """ヘッダー行を書き込む。"""
    for col_idx, (col_name, width) in enumerate(zip(columns, widths), 1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["header_border"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_cell(ws, row, col, value, styles, font_key="data_font",
                fill_key="bg_fill", num_fmt=None, align="right"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = styles[fill_key]
    cell.font = styles[font_key]
    cell.border = styles["border"]
    if align:
        cell.alignment = Alignment(horizontal=align)
    if num_fmt and value is not None:
        cell.number_format = num_fmt
    return cell


def write_sheet1_sentiment(wb, token_results, changes, styles):
    """Sheet 1: Token Sentiment"""
    ws = wb.active
    ws.title = "Token Sentiment"

    columns = [
        "Rank", "Token", "LONG人数", "SHORT人数", "合計",
        "LONG%", "SHORT%", "Sentiment",
        "LONG価値($)", "SHORT価値($)", "合計価値($)", "Net方向($)",
        "LONG PnL($)", "SHORT PnL($)", "合計PnL($)",
        "平均Lev(L/S)",
        "LONG%変化(pt)", "価値変動率(%)", "トレーダー数変化",
    ]
    widths = [6, 12, 10, 10, 6, 8, 8, 18,
              16, 16, 16, 14, 14, 14, 14, 14, 14, 14, 14]

    _apply_header(ws, 1, columns, widths, styles)

    for i, t in enumerate(token_results):
        row = i + 2
        coin = t["coin"]
        ch = changes.get(coin, {})

        # LONG%/SHORT% は整数パーセント
        long_pct = round(t["long_ratio"] * 100)
        short_pct = round(t["short_ratio"] * 100)

        values = [
            (i + 1, None, "right"),
            (coin, "data_font", "left"),
            (t["long_count"], "long_font", "right"),
            (t["short_count"], "short_font", "right"),
            (t["total_count"], None, "right"),
            (long_pct, None, "right"),
            (short_pct, None, "right"),
            (t["sentiment"], None, "left"),
            (t["long_value"], None, "right"),
            (t["short_value"], None, "right"),
            (t["total_value"], None, "right"),
            (t["net_direction"], None, "right"),
            (t["long_pnl"], None, "right"),
            (t["short_pnl"], None, "right"),
            (t["total_pnl"], None, "right"),
            (f'{t["avg_lev_long"]:.1f}/{t["avg_lev_short"]:.1f}', None, "center"),
            (ch.get("long_ratio_change"), None, "right"),
            (ch.get("value_change_pct"), None, "right"),
            (ch.get("trader_count_change"), None, "right"),
        ]

        for col_idx, (val, font_override, align) in enumerate(values, 1):
            col_name = columns[col_idx - 1]

            # フォント選択
            if font_override:
                font_key = font_override
            elif col_name == "Net方向($)" and val is not None:
                font_key = "pnl_pos_font" if val >= 0 else "pnl_neg_font"
            elif "PnL" in col_name and val is not None:
                font_key = "pnl_pos_font" if val >= 0 else "pnl_neg_font"
            elif col_name == "LONG%変化(pt)" and val is not None:
                font_key = "pnl_pos_font" if val >= 0 else "pnl_neg_font"
            elif col_name == "価値変動率(%)" and val is not None:
                font_key = "pnl_pos_font" if val >= 0 else "pnl_neg_font"
            else:
                font_key = "data_font"

            # 数値フォーマット
            num_fmt = None
            if isinstance(val, float):
                if col_name in ("LONG%変化(pt)", "価値変動率(%)"):
                    num_fmt = "+0.0;-0.0;0"
                else:
                    num_fmt = "#,##0"
            elif isinstance(val, int) and col_name in ("LONG価値($)", "SHORT価値($)",
                                                        "合計価値($)"):
                num_fmt = "#,##0"

            _apply_cell(ws, row, col_idx, val, styles,
                        font_key=font_key, num_fmt=num_fmt, align=align)

    ws.freeze_panes = "A2"


def write_sheet2_chart(wb, token_results, styles):
    """Sheet 2: Top30 Chart — 7列データテーブル + LONG vs SHORT 積み上げ棒グラフ"""
    ws = wb.create_sheet("Top30 Chart")
    top30 = token_results[:30]

    # 7列ヘッダー
    headers = ["Rank", "Token", "LONG人数", "SHORT人数",
               "LONG価値($M)", "SHORT価値($M)", "LONG%"]
    widths = [6, 12, 10, 10, 14, 14, 8]
    _apply_header(ws, 1, headers, widths, styles)

    for i, t in enumerate(top30):
        row = i + 2
        long_pct = round(t["long_ratio"] * 100)
        _apply_cell(ws, row, 1, i + 1, styles, align="right")
        _apply_cell(ws, row, 2, t["coin"], styles, align="left")
        _apply_cell(ws, row, 3, t["long_count"], styles,
                    font_key="long_font", align="right")
        _apply_cell(ws, row, 4, t["short_count"], styles,
                    font_key="short_font", align="right")
        _apply_cell(ws, row, 5, round(t["long_value"] / 1_000_000, 2), styles,
                    font_key="long_font", num_fmt="#,##0.00", align="right")
        _apply_cell(ws, row, 6, round(t["short_value"] / 1_000_000, 2), styles,
                    font_key="short_font", num_fmt="#,##0.00", align="right")
        _apply_cell(ws, row, 7, long_pct, styles, align="right")

    # 棒グラフ作成
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "Top 30 Tokens: LONG vs SHORT Value ($M)"
    chart.y_axis.title = "Position Value ($M)"
    chart.x_axis.title = "Token"
    chart.width = 30
    chart.height = 15
    chart.style = 10

    data_end = min(len(top30) + 1, 31)
    cats = Reference(ws, min_col=2, min_row=2, max_row=data_end)
    long_data = Reference(ws, min_col=5, min_row=1, max_row=data_end)
    short_data = Reference(ws, min_col=6, min_row=1, max_row=data_end)

    chart.add_data(long_data, titles_from_data=True)
    chart.add_data(short_data, titles_from_data=True)
    chart.set_categories(cats)

    # LONG=緑, SHORT=赤
    chart.series[0].graphicalProperties.solidFill = STYLES["long_color"].lstrip("#")
    chart.series[1].graphicalProperties.solidFill = STYLES["short_color"].lstrip("#")

    ws.add_chart(chart, "I2")


def write_sheet3_matrix(wb, rows, token_results, styles):
    """Sheet 3: Trader×Token Matrix
    列: Tier, Trader, AcctVal($), [上位30トークン...]
    セル: L(LONG) / S(SHORT) / 空欄(ポジションなし)
    """
    ws = wb.create_sheet("Trader\u00d7Token Matrix")

    top30_coins = [t["coin"] for t in token_results[:30]]
    active_traders, matrix = build_trader_matrix(rows, top30_coins)

    # ヘッダー: 先頭3列 + トークン列
    fixed_headers = ["Tier", "Trader", "AcctVal($)"]
    fixed_widths = [6, 22, 14]
    for col_idx, (h, w) in enumerate(zip(fixed_headers, fixed_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["header_border"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for col_idx, coin in enumerate(top30_coins, 4):
        cell = ws.cell(row=1, column=col_idx, value=coin)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["header_border"]
        cell.alignment = Alignment(horizontal="center", text_rotation=90)
        ws.column_dimensions[get_column_letter(col_idx)].width = 5

    # データ行
    for row_idx, trader_info in enumerate(active_traders, 2):
        trader = trader_info["trader"]
        _apply_cell(ws, row_idx, 1, trader_info["tier"], styles,
                    font_key="accent_font", align="center")
        _apply_cell(ws, row_idx, 2, trader, styles,
                    font_key="data_font", align="left")
        _apply_cell(ws, row_idx, 3, trader_info["acct_value"], styles,
                    num_fmt="#,##0", align="right")

        for col_idx, coin in enumerate(top30_coins, 4):
            val = matrix.get((trader, coin))
            if val == "L":
                _apply_cell(ws, row_idx, col_idx, "L", styles,
                            font_key="long_font", fill_key="long_row_fill",
                            align="center")
            elif val == "S":
                _apply_cell(ws, row_idx, col_idx, "S", styles,
                            font_key="short_font", fill_key="short_row_fill",
                            align="center")
            else:
                # ポジションなし: 全角ダッシュ
                _apply_cell(ws, row_idx, col_idx, "\u2015", styles, align="center")

    ws.freeze_panes = "D2"


def write_sheet4_notes(wb, timestamp, prev_file, token_results, rows, styles):
    """Sheet 4: 凡例・注記（サンプル準拠フォーマット）"""
    ws = wb.create_sheet("\u51e1\u4f8b\u30fb\u6ce8\u8a18")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    total_traders = len(TRADERS)
    active_count = len(set(r["trader"] for r in rows))
    total_positions = len(rows)
    token_count = len(token_results)

    notes = [
        ("\U0001f4c5 \u30b9\u30ca\u30c3\u30d7\u30b7\u30e7\u30c3\u30c8\u6642\u523b",
         timestamp),
        ("\U0001f4c2 \u6bd4\u8f03\u5bfe\u8c61\u30d5\u30a1\u30a4\u30eb",
         prev_file if prev_file else "\u306a\u3057\uff08\u521d\u56de\u5b9f\u884c\uff09"),
        ("", ""),
        ("\u30bb\u30f3\u30c1\u30e1\u30f3\u30c8\u5224\u5b9a\u57fa\u6e96", ""),
        ("\U0001f7e2 Strong LONG", "LONG\u6bd4\u7387 \u2265 70%"),
        ("\u2197 Lean LONG", "LONG\u6bd4\u7387 55\u201369%"),
        ("\u2696 Neutral", "LONG\u6bd4\u7387 = 50%"),
        ("\u2198 Lean SHORT", "SHORT\u6bd4\u7387 55\u201369%"),
        ("\U0001f534 Strong SHORT", "SHORT\u6bd4\u7387 \u2265 70%"),
        ("", ""),
        ("\u30c7\u30fc\u30bf\u30bd\u30fc\u30b9",
         "Hyperliquid API /info clearinghouseState"),
        ("\u5bfe\u8c61\u30c8\u30ec\u30fc\u30c0\u30fc\u6570",
         f"{total_traders}\u540d\uff08T1\u2013T4 tier\uff09"),
        ("\u30a2\u30af\u30c6\u30a3\u30d6\u30c8\u30ec\u30fc\u30c0\u30fc\u6570",
         f"{active_count}\u540d"),
        ("\u53d6\u5f97\u30c8\u30fc\u30af\u30f3\u6570",
         f"{token_count} tokens\uff08\u30dd\u30b8\u30b7\u30e7\u30f3\u3042\u308a\uff09"),
        ("\u7dcf\u30dd\u30b8\u30b7\u30e7\u30f3\u6570", str(total_positions)),
        ("Matrix\u8868\u793a\u5217",
         "\u4e0a\u4f4d30\u30c8\u30fc\u30af\u30f3\uff08\u5408\u8a08\u4fa1\u5024\u9806\uff09"),
        ("", ""),
        ("Matrix \u8272\u51e1\u4f8b",
         "L=LONG(\u7dd1) S=SHORT(\u8d64) \u2015=\u30dd\u30b8\u30b7\u30e7\u30f3\u306a\u3057"),
    ]

    for row_idx, (label, value) in enumerate(notes, 1):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.fill = styles["bg_fill"]
        label_cell.border = styles["border"]
        if label.startswith(("\u30bb\u30f3\u30c1\u30e1\u30f3\u30c8", "Matrix")):
            label_cell.font = styles["accent_font"]
        elif label.startswith(("\U0001f4c5", "\U0001f4c2")):
            label_cell.font = styles["accent_font"]
        else:
            label_cell.font = styles["header_font"]

        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.fill = styles["bg_fill"]
        value_cell.font = styles["data_font"]
        value_cell.border = styles["border"]


# ============================================================
# メインエントリポイント
# ============================================================

def analyze_sentiment(positions_file: str, logger=None) -> str | None:
    """ポジションExcelを入力としてセンチメント分析を実行。

    Args:
        positions_file: Phase 1 で生成されたExcelファイルパス
        logger: ロガー

    Returns: 分析結果Excelファイルパス。失敗時None。
    """
    if logger is None:
        logger = setup_logging()

    logger.info(f"Phase 2: Loading positions from {Path(positions_file).name}")

    # ポジション読み込み
    rows = load_positions(positions_file)
    if not rows:
        logger.error("No position data found. Skipping analysis.")
        return None

    # トークン別集計
    token_results = analyze_tokens(rows)
    logger.info(f"Phase 2: Analyzing sentiment ({len(token_results)} tokens)")

    # 前回比較
    prev_file = find_previous_file(positions_file)
    changes = {}
    if prev_file:
        logger.info(f"Comparing with: {Path(prev_file).name}")
        prev_rows = load_positions(prev_file)
        if prev_rows:
            prev_tokens = analyze_tokens(prev_rows)
            changes = compute_changes(token_results, prev_tokens)
    else:
        logger.info("No previous file found. Skipping comparison.")

    # タイムスタンプ抽出（ファイル名から）
    basename = Path(positions_file).stem  # hl_positions_YYYY-MM-DD_HH-mm
    ts_part = basename.replace("hl_positions_", "")
    output_file = DATA_DIR / f"hl_analysis_{ts_part}.xlsx"

    # Excel出力
    styles = build_styles()
    wb = Workbook()

    write_sheet1_sentiment(wb, token_results, changes, styles)
    write_sheet2_chart(wb, token_results, styles)
    write_sheet3_matrix(wb, rows, token_results, styles)

    prev_name = Path(prev_file).name if prev_file else None
    write_sheet4_notes(wb, ts_part, prev_name, token_results, rows, styles)

    wb.save(output_file)
    logger.info(f"Saved: {output_file.relative_to(DATA_DIR.parent)}")
    return str(output_file)


if __name__ == "__main__":
    log = setup_logging()

    # 引数でファイル指定 or data/ 内の最新を使用
    if len(sys.argv) > 1:
        pos_file = sys.argv[1]
    else:
        pattern = str(DATA_DIR / "hl_positions_*.xlsx")
        files = sorted(glob.glob(pattern))
        if not files:
            log.error("No position files found in data/")
            sys.exit(1)
        pos_file = files[-1]
        log.info(f"Using latest: {Path(pos_file).name}")

    result = analyze_sentiment(pos_file, log)
    if result:
        log.info("=== Phase 2 Done ===")
    else:
        log.error("=== Phase 2 Failed ===")

"""Hyperliquid Position Tracker — 設定定数・トレーダーリスト・ロギング・共通スタイル"""

import logging
import os
from datetime import datetime
from pathlib import Path

from openpyxl.styles import Border, Font, PatternFill, Side

# ============================================================
# ディレクトリパス
# ============================================================
BASE_DIR = Path(__file__).resolve().parent.parent
SCRIPTS_DIR = BASE_DIR / "scripts"
DATA_DIR = BASE_DIR / "data"
LOGS_DIR = BASE_DIR / "logs"

# ============================================================
# API設定
# ============================================================
API_URL = "https://api.hyperliquid.xyz/info"
API_TIMEOUT = 15  # seconds
API_RETRY_DELAY = 3  # seconds
API_RATE_LIMIT = 1  # seconds between requests

# ============================================================
# ChromaDB設定
# ============================================================
CHROMADB_DIR = BASE_DIR / "chromadb_data"
CHROMADB_COLLECTION_SENTIMENT = "token_sentiment"
CHROMADB_COLLECTION_MATRIX = "trader_token_matrix"
# 保持期間を超えたスナップショットは Phase 3 の末尾で削除される
CHROMADB_RETENTION_DAYS = 30

# ============================================================
# トレーダーリスト（26名・4 Tier）
# (Tier, Rank, Name, Address, 備考)
# ============================================================
TRADERS = [
    # --- T1: 機関・有名 ---
    ("T1", 1, "BobbyBigSize", "0x7fdafde5cfb5465924316eced2d3715494c517d1", "Fasanara Capital関連。ショート専門"),
    ("T1", 8, "x.com/SilkBtc", "0x880ac484a1743862989a441d6d867238c7aa311c", "X(Twitter)アカウント公開済"),
    ("T1", 17, "Auros", "0x023a3d058020fb76cca98f01b3c48c8938a22355", "機関マーケットメイカー。取引高$225B"),
    ("T1", 38, "ABC", "0x162cc7c861ebd0c06b3d72319201150482518185", "アルゴトレーダー。取引高$55.8B"),
    # --- T2: Jeff系グループ ---
    ("T2", 4, "thank you jefef", "0xfae95f601f3a25ace60d19dbb929f2a5c57e3571", "Jeff系グループ（合算$316M+）"),
    ("T2", 15, "jefe", "0x51156f7002c4f74f4956c9e0f2b7bfb6e9dbfac2", "Jeff系グループ"),
    ("T2", 43, "fuck jeff gib S3", "0xa464abbf049fb75585484addcbc00169062e813a", "Jeff系グループ"),
    ("T2", 58, "NMTD -Thank you Jeff", "0xf517639a8872e756ac98d3c65507d2ebc25cc032", "Jeff系グループ"),
    ("T2", 79, "thank you jeffff", "0x5c02f2dfcb6537b83929596fe8a3278e237e3e7c", "Jeff系グループ"),
    ("T2", 88, "thank you JEFF", "0x2d23b731e5f04996a2dfdbe434c7d922afdb5e00", "Jeff系グループ"),
    # --- T3: コミュニティ注目 ---
    ("T3", 20, "憨巴小龙", "0x8e096995c3e4a3f0bc5b3ea1cba94de2aa4d70c9", "中国語圏の大物"),
    ("T3", 59, "韦小宝.eth", "0x99b1098d9d50aa076f78bd26ab22e6abd3710729", "ENSユーザー"),
    ("T3", 61, "SSS888", "0x598f9efb3164ec216b4eff33c2b239605be5af8e", "残高$16→PNL $24M"),
    ("T3", 64, "wanyekest", "0x8607a7d180de23645db594d90621d837749408d5", "$13.3M安定残高"),
    ("T3", 65, "Prison Su Zhu", "0xdafc555a97b358bc09fccf4c0e583c1ec5838b16", "Su Zhuパロディ名"),
    ("T3", 67, "EARLY", "0x1da722cfa8b2dfda57cf8d787689039c7a63f049", "HYPE初期参入者"),
    ("T3", 68, "Penision Fund", "0x0ddf9bae2af4b874b96d287a5ad42eb47138a902", "$32.6M安定残高"),
    ("T3", 76, "demo_account", "0x34971bc50eb4484505e4a24516c8db843fbef162", ""),
    ("T3", 80, "VidaBWE", "0x71dfc07de32c2ebf1c4801f4b1c9e40b76d4a23d", "$15M安定残高"),
    ("T3", 96, "jez", "0xaa7577a7a27aa7fcf6d0ec481b87df3ad0f6a88e", ""),
    # --- T4: 匿名高PNL ---
    ("T4", 1, "0xecb6...2b00", "0xecb63caa47c7c4e77f60f1ce858cf28dc2b82b00", "全期間PNL 1位"),
    ("T4", 2, "0x5b5d...c060", "0x5b5d51203a0f9079f8aeb098a6523a13f298c060", ""),
    ("T4", 5, "0x9794...333b", "0x9794bbbc222b6b93c1417d01aa1ff06d42e5333b", "残高$1,395→PNL $144M"),
    ("T4", 6, "0x20c2...44f5", "0x20c2d95a3dfdca9e9ad12794d5fa6fad99da44f5", ""),
    ("T4", 9, "0x2ea1...23f4", "0x2ea18c23f72a4b6172c55b411823cdc5335923f4", "残高$1.99"),
    ("T4", 10, "0xbde2...60b1", "0xbde2ddc49a2e6827300faa6afc93d572114a60b1", "残高$1.01。ROI最高"),
]

# ============================================================
# Excelスタイリング定数（ダークテーマ）
# ============================================================
STYLES = {
    "bg": "#0D1117",
    "header_bg": "#1B2A3D",
    "header_font_color": "#FFFFFF",
    "data_font_color": "#CCCCCC",
    "long_color": "#00FF88",
    "short_color": "#FF6666",
    "accent_color": "#FFD700",
    "pnl_positive": "#00FF88",
    "pnl_negative": "#FF6666",
    "long_row_bg": "#0A2E1A",
    "short_row_bg": "#2E0A0A",
    "border_color": "#333333",
    "header_underline": "#00D4AA",
}


# ============================================================
# Excelスタイル構築（共通）
# ============================================================

def hex_to_argb(hex_color: str) -> str:
    """#RRGGBB → FFRRGGBB (openpyxl用)"""
    return "FF" + hex_color.lstrip("#")


def build_styles() -> dict:
    """openpyxlスタイルオブジェクトを構築して返す。"""
    s = STYLES
    border = Border(
        left=Side(style="thin", color=hex_to_argb(s["border_color"])),
        right=Side(style="thin", color=hex_to_argb(s["border_color"])),
        top=Side(style="thin", color=hex_to_argb(s["border_color"])),
        bottom=Side(style="thin", color=hex_to_argb(s["border_color"])),
    )
    header_border = Border(
        left=Side(style="thin", color=hex_to_argb(s["border_color"])),
        right=Side(style="thin", color=hex_to_argb(s["border_color"])),
        top=Side(style="thin", color=hex_to_argb(s["border_color"])),
        bottom=Side(style="medium", color=hex_to_argb(s["header_underline"])),
    )
    return {
        "bg_fill": PatternFill("solid", fgColor=hex_to_argb(s["bg"])),
        "header_fill": PatternFill("solid", fgColor=hex_to_argb(s["header_bg"])),
        "header_font": Font(name="Arial", size=10, bold=True,
                            color=hex_to_argb(s["header_font_color"])),
        "data_font": Font(name="Arial", size=9,
                          color=hex_to_argb(s["data_font_color"])),
        "long_font": Font(name="Arial", size=9,
                          color=hex_to_argb(s["long_color"])),
        "short_font": Font(name="Arial", size=9,
                           color=hex_to_argb(s["short_color"])),
        "pnl_pos_font": Font(name="Arial", size=9,
                             color=hex_to_argb(s["pnl_positive"])),
        "pnl_neg_font": Font(name="Arial", size=9,
                             color=hex_to_argb(s["pnl_negative"])),
        "accent_font": Font(name="Arial", size=9, bold=True,
                            color=hex_to_argb(s["accent_color"])),
        "long_row_fill": PatternFill("solid", fgColor=hex_to_argb(s["long_row_bg"])),
        "short_row_fill": PatternFill("solid", fgColor=hex_to_argb(s["short_row_bg"])),
        "border": border,
        "header_border": header_border,
    }


# ============================================================
# ロギング設定
# ============================================================
def setup_logging() -> logging.Logger:
    """ファイル+コンソール出力のロガーをセットアップして返す。

    ログファイル: logs/hl_YYYY-MM-DD_HH-mm.log
    フォーマット: YYYY-MM-DD HH:MM:SS [LEVEL] メッセージ
    """
    LOGS_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    log_file = LOGS_DIR / f"hl_{timestamp}.log"

    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    datefmt = "%Y-%m-%d %H:%M:%S"

    logger = logging.getLogger("hl_tracker")
    logger.setLevel(logging.DEBUG)

    # 既存ハンドラがあればクリア（再呼び出し時の重複防止）
    if logger.handlers:
        logger.handlers.clear()

    # ファイルハンドラ
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(fmt, datefmt=datefmt))
    logger.addHandler(fh)

    # コンソールハンドラ
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter(fmt, datefmt=datefmt))
    logger.addHandler(ch)

    return logger

"""Phase 1: Hyperliquid トレーダーポジション取得 & Excel出力"""

from __future__ import annotations

import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

# スタンドアロン実行時のパス解決
if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import logging
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from scripts.config import (
    API_RATE_LIMIT,
    API_RETRY_DELAY,
    API_TIMEOUT,
    API_URL,
    DATA_DIR,
    TRADERS,
    build_styles,
    setup_logging,
)

# Excel列定義
COLUMNS = [
    "Tier",
    "Trader",
    "Acct Value($)",
    "Coin",
    "Side",
    "Size",
    "Leverage",
    "Entry Px",
    "Pos Value($)",
    "Unrealized PnL($)",
    "ROE(%)",
    "Liq Px",
    "Margin($)",
    "Cum Funding($)",
]

# 列幅設定
COL_WIDTHS = [6, 22, 16, 10, 7, 14, 9, 14, 16, 18, 10, 14, 14, 16]


def fetch_trader_positions(address: str, logger) -> dict | None:
    """1トレーダーのポジションをAPI取得。失敗時は1回リトライ。"""
    payload = {"type": "clearinghouseState", "user": address}
    headers = {"Content-Type": "application/json"}

    for attempt in range(2):
        try:
            resp = requests.post(
                API_URL, json=payload, headers=headers, timeout=API_TIMEOUT
            )
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            if attempt == 0:
                logger.warning(f"  API error for {address[:10]}..., retrying: {e}")
                time.sleep(API_RETRY_DELAY)
            else:
                logger.error(f"  Failed after retry for {address[:10]}...: {e}")
    return None


def parse_positions(data: dict, logger=None) -> tuple[float, list[dict]]:
    """APIレスポンスからアカウント残高とポジションリストをパース。"""
    _log = logger or logging.getLogger("hl_tracker")
    acct_value = float(data.get("marginSummary", {}).get("accountValue", 0))

    positions = []
    for ap in data.get("assetPositions", []):
        pos = ap.get("position", {})
        try:
            szi = float(pos.get("szi", 0))
            if szi == 0:
                continue

            # Fix 6: leverage がオブジェクトでない場合のハンドリング
            leverage_info = pos.get("leverage", 0)
            if isinstance(leverage_info, dict):
                leverage = float(leverage_info.get("value", 0))
            else:
                try:
                    leverage = float(leverage_info) if leverage_info else 0
                except (ValueError, TypeError):
                    leverage = 0

            liq_px_raw = pos.get("liquidationPx")
            liq_px = float(liq_px_raw) if liq_px_raw is not None else None

            positions.append({
                "coin": pos.get("coin", ""),
                "side": "LONG" if szi > 0 else "SHORT",
                "size": abs(szi),
                "leverage": leverage,
                "entry_px": float(pos.get("entryPx", 0)),
                "pos_value": float(pos.get("positionValue", 0)),
                "unrealized_pnl": float(pos.get("unrealizedPnl", 0)),
                "roe": float(pos.get("returnOnEquity", 0)) * 100,
                "liq_px": liq_px,
                "margin": float(pos.get("marginUsed", 0)),
                "cum_funding": float(pos.get("cumFunding", {}).get("sinceOpen", 0)),
            })
        except (ValueError, TypeError, KeyError) as e:
            coin = pos.get("coin", "unknown")
            _log.warning(f"  Skipping position {coin}: parse error: {e}")
            continue
    return acct_value, positions


def write_excel_file(all_data: list[dict], timestamp: str, logger) -> str:
    """ポジションデータをExcelに書き出す。返り値はファイルパス。"""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"hl_positions_{timestamp}.xlsx"
    filepath = DATA_DIR / filename

    wb = Workbook()
    ws = wb.active
    date_str = timestamp.split("_")[0]
    ws.title = f"Positions {date_str}"

    styles = build_styles()

    # ヘッダー行
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["header_border"]
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 列幅設定
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # データ行
    row = 2
    for trader_data in all_data:
        tier = trader_data["tier"]
        name = trader_data["name"]
        acct_value = trader_data["acct_value"]
        positions = trader_data["positions"]

        if not positions:
            # ポジションなし
            _write_row(ws, row, styles, tier, name, acct_value,
                       "(No positions)", "", None, None, None, None,
                       None, None, None, None, None, is_first=True)
            row += 1
            continue

        for i, pos in enumerate(positions):
            is_first = (i == 0)
            _write_row(
                ws, row, styles,
                tier if is_first else "",
                name if is_first else "",
                acct_value if is_first else None,
                pos["coin"], pos["side"], pos["size"], pos["leverage"],
                pos["entry_px"], pos["pos_value"], pos["unrealized_pnl"],
                pos["roe"], pos["liq_px"], pos["margin"], pos["cum_funding"],
                is_first=is_first,
            )
            row += 1

    # フリーズペイン（ヘッダー固定）
    ws.freeze_panes = "A2"

    wb.save(filepath)
    logger.info(f"Saved: {filepath.relative_to(DATA_DIR.parent)}")
    return str(filepath)


def _write_row(ws, row, styles, tier, name, acct_value,
               coin, side, size, leverage, entry_px, pos_value,
               unrealized_pnl, roe, liq_px, margin, cum_funding,
               is_first=False):
    """1行分のデータをワークシートに書き込む。"""
    # 行背景色の決定
    if side == "LONG":
        row_fill = styles["long_row_fill"]
    elif side == "SHORT":
        row_fill = styles["short_row_fill"]
    else:
        row_fill = styles["bg_fill"]

    values = [
        tier, name, acct_value, coin, side, size, leverage,
        entry_px, pos_value, unrealized_pnl, roe, liq_px, margin, cum_funding,
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = row_fill
        cell.border = styles["border"]

        # フォント決定
        col_name = COLUMNS[col_idx - 1]
        if col_name == "Tier" and val:
            cell.font = styles["accent_font"]
        elif col_name == "Side":
            cell.font = styles["long_font"] if val == "LONG" else styles["short_font"]
        elif col_name in ("Unrealized PnL($)", "ROE(%)") and val is not None:
            cell.font = styles["pnl_pos_font"] if val >= 0 else styles["pnl_neg_font"]
        else:
            cell.font = styles["data_font"]

        # 数値列は右寄せ
        if col_idx >= 3 and val is not None:
            cell.alignment = Alignment(horizontal="right")
            if isinstance(val, float):
                if col_name in ("Entry Px", "Liq Px"):
                    cell.number_format = "#,##0.00"
                elif col_name == "ROE(%)":
                    cell.number_format = "#,##0.00"
                elif col_name == "Size":
                    cell.number_format = "#,##0.#####"
                else:
                    cell.number_format = "#,##0.00"


def fetch_positions(logger=None, write_excel=False) -> tuple[list[dict], str] | None:
    """全トレーダーのポジションを取得。オプションでExcel出力。

    Args:
        logger: ロガー
        write_excel: TrueならExcelファイルを生成

    Returns: (all_data, timestamp) のタプル。全失敗時はNone。
    """
    if logger is None:
        logger = setup_logging()

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    logger.info(f"Phase 1: Fetching positions for {len(TRADERS)} traders")

    all_data = []
    success_count = 0
    total_positions = 0

    for idx, (tier, rank, name, address, note) in enumerate(TRADERS):
        data = fetch_trader_positions(address, logger)

        if data is None:
            logger.error(f"  [{tier}] {name}: Skipped (API failure)")
            continue

        acct_value, positions = parse_positions(data)
        pos_count = len(positions)
        total_positions += pos_count
        success_count += 1

        acct_str = f"${acct_value:,.0f}" if acct_value else "$0"
        logger.info(f"  [{tier}] {name}: {pos_count} positions (Acct: {acct_str})")

        all_data.append({
            "tier": tier,
            "rank": rank,
            "name": name,
            "address": address,
            "acct_value": acct_value,
            "positions": positions,
        })

        # レート制限（最後のリクエスト以外）
        if idx < len(TRADERS) - 1:
            time.sleep(API_RATE_LIMIT)

    if success_count == 0:
        logger.error("All traders failed.")
        return None

    active_traders = sum(1 for d in all_data if d["positions"])
    logger.info(
        f"Phase 1 complete: {total_positions} positions "
        f"from {active_traders} active traders"
    )

    if write_excel:
        write_excel_file(all_data, timestamp, logger)

    return all_data, timestamp


if __name__ == "__main__":
    log = setup_logging()
    excel_flag = "--excel" in sys.argv
    result = fetch_positions(log, write_excel=excel_flag)
    if result:
        log.info("=== Phase 1 Done ===")
    else:
        log.error("=== Phase 1 Failed ===")
